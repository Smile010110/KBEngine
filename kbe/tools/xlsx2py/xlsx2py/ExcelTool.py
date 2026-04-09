# -*- coding: utf-8 -*-

"""
Excel 操作封装（基于 openpyxl）
"""

from __future__ import annotations

import os
from typing import Any, Optional

import openpyxl
from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet


class ExcelTool:
    """
    简单封装 Excel 各种操作
    当前依赖: openpyxl
    """

    def __init__(self, file_name: str):
        self.fileName = os.path.abspath(file_name)
        self.__workbook: Optional[Workbook] = None

    def getWorkbookEx(self, auto_create: bool = False) -> bool:
        try:
            if os.path.exists(self.fileName):
                self.__workbook = openpyxl.load_workbook(self.fileName)
            elif auto_create:
                self.__workbook = openpyxl.Workbook()
            else:
                return False
            return True
        except Exception:
            if auto_create:
                self.__workbook = openpyxl.Workbook()
                return True
            return False

    def getXLSX(self) -> Workbook:
        if self.__workbook is None:
            raise RuntimeError("Workbook has not been opened.")
        return self.__workbook

    def close(self, saveChanges: bool = False) -> None:
        if self.__workbook is None:
            return

        if saveChanges:
            self.__workbook.save(self.fileName)

        self.__workbook.close()
        self.__workbook = None

    def save(self) -> None:
        if self.__workbook is None:
            raise RuntimeError("Workbook has not been opened.")
        self.__workbook.save(self.fileName)

    def getSheetCount(self) -> int:
        if self.__workbook is None:
            return 0
        return len(self.__workbook.worksheets)

    def getSheetNameByIndex(self, index: int) -> str:
        return self.getXLSX().worksheets[index].title

    def getSheetByIndex(self, index: int) -> Optional[Worksheet]:
        try:
            return self.getXLSX().worksheets[index]
        except Exception:
            return None

    def getSheetByName(self, name: str) -> Optional[Worksheet]:
        wb = self.getXLSX()
        if name in wb.sheetnames:
            return wb[name]
        return None

    def createSheet(self, title: str) -> Worksheet:
        wb = self.getXLSX()
        return wb.create_sheet(title=title)

    def removeSheet(self, sheet: Worksheet) -> None:
        wb = self.getXLSX()
        wb.remove(sheet)

    def renameSheet(self, sheet: Worksheet, title: str) -> None:
        sheet.title = title

    def __getRowCountOnSheet(self, sheet: Worksheet) -> int:
        """
        获得有数据的最大列数（按原工程命名习惯保留 rowCount）
        """
        max_col = sheet.max_column or 0
        max_row = sheet.max_row or 0

        for col in range(max_col, 0, -1):
            has_data = False
            for row in range(1, max_row + 1):
                if sheet.cell(row=row, column=col).value not in (None, ""):
                    has_data = True
                    break
            if has_data:
                return col
        return 0

    def getRowCount(self, sheetIndex: int) -> int:
        """
        获得一排有多少元素（实际返回有效列数）
        """
        ws = self.getXLSX().worksheets[sheetIndex]
        return self.__getRowCountOnSheet(ws)

    def getColCount(self, sheetIndex: int) -> int:
        """
        获得一列有多少元素（实际返回有效行数）
        """
        ws = self.getXLSX().worksheets[sheetIndex]
        max_row = ws.max_row or 0
        max_col = ws.max_column or 0

        for row in range(max_row, 0, -1):
            has_data = False
            for col in range(1, max_col + 1):
                if ws.cell(row=row, column=col).value not in (None, ""):
                    has_data = True
                    break
            if has_data:
                return row
        return 0

    def getValue(self, sheet: Worksheet, row: int, col: int) -> Any:
        return sheet.cell(row=row, column=col).value

    def getText(self, sheet: Worksheet, row: int, col: int) -> str:
        value = sheet.cell(row=row, column=col).value
        return "" if value is None else str(value)

    def getRowValues(self, sheet: Worksheet, row: int) -> list[Any]:
        """
        兼容老代码：row 为 0-based
        """
        cc = self.__getRowCountOnSheet(sheet)
        excel_row = row + 1
        return [sheet.cell(row=excel_row, column=i).value for i in range(1, cc + 1)]

    def getColValues(self, sheet: Worksheet, col: int) -> list[Any]:
        """
        兼容老代码：col 为 0-based
        """
        rc = self.getColCountBySheet(sheet)
        excel_col = col + 1
        return [sheet.cell(row=i, column=excel_col).value for i in range(1, rc + 1)]

    def getColCountBySheet(self, sheet: Worksheet) -> int:
        max_row = sheet.max_row or 0
        max_col = sheet.max_column or 0

        for row in range(max_row, 0, -1):
            has_data = False
            for col in range(1, max_col + 1):
                if sheet.cell(row=row, column=col).value not in (None, ""):
                    has_data = True
                    break
            if has_data:
                return row
        return 0
